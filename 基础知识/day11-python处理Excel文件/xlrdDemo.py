#!/usr/bin/python
# -*- coding: UTF-8 -*-
import openpyxl as openpyxl

def demo1():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    # 全部sheet对象
    worksheets = workbook.worksheets
    print("全部sheet对象:",worksheets)
    # 全部sheet名称
    sheet_name = workbook.sheetnames
    print("全部sheet名称:",sheet_name)
    # 按名称读取sheet
    sheet_1 = workbook["SheetJS"]
    print("按名称读取sheet:",sheet_1)
    # 按下标读取
    sheet_2 = worksheets[0]
    print("按下标读取sheet:", sheet_2)
    # 获取当前正在使用的sheet
    sheet_active= workbook.active
    print("获取当前正在使用的sheet:",sheet_active)
    # 获取sheet的属性
    sheet_title = sheet_active.title
    print("获取sheet的属性:",sheet_title)
    # 获取行
    max_row = sheet_active.max_row
    print("获取行:",max_row)
    # 获取列
    max_column = sheet_active.max_column
    print("获取列:",max_column)


def demo2():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    print(sheet)
    workbook.create_sheet("我是一个新的sheet")
    print(workbook.sheetnames)
    workbook.save(file_path)

def demo3():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    # 修改工作表名称
    workbook.active.title = "test"
    # 保存文件
    workbook.save(file_path)

def demo4():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet_active = workbook.active
    # remove删除工作表
    # workbook.remove(sheet_active)
    # del操作删除
    del workbook["test"]
    # 保存文件
    workbook.save(file_path)

def demo5():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # 遍历全部行
    for row in sheet.rows:
        print("row:",row)
    # 读取部分行列
    print("-----------------------")
    print(sheet[1])
    print("-----------------------")
    print(sheet["A:B"])

def demo6():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # 在第4行插入1行空行
    sheet.insert_rows(4)
    # 在第2行插入2行空行
    sheet.insert_rows(idx=2,amount=2)
    # 添加一行数据到表
    row_data = ["tom", 15, "tom@test.com"]
    sheet.append(row_data)
    workbook.save(file_path)


def demo7():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # 删除行
    sheet.delete_rows(2, 2)
    workbook.save(file_path)


def demo8():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # 指定行列坐标获取单元格
    cell_1 = sheet["A1"]
    print("指定行列坐标获取单元格:",cell_1)

    # cell函数获取单元格
    cell_2 = sheet.cell(row=1,column=1)
    print("cell函数获取单元格:",cell_2)

    # cell 值
    value = cell_2.value
    print("cell 值:",value)
    print("cell_2.coordinate:",cell_2.coordinate)
    print("cell_2.row:",cell_2.row)
    print("cell_2.column:",cell_2.column)


def demo9():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # 移动数据区域(向上移动2行，向右移动3列)，正整数为向下或向右，负整数为向上或向左
    sheet.move_range("A3:C3", rows=-2, cols=3)
    workbook.save(file_path)

def demo10():
    file_path = "/Users/yunna/Desktop/1.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # 单元格合并，使用范围坐标
    sheet.merge_cells("A2:B3")
    # 单元格合并，指定行列下标（下标从1开始）
    sheet.merge_cells(start_row=5, start_column=3, end_row=7, end_column=4)
    workbook.save(file_path)
    # 拆分单元格
    sheet.unmerge_cells("A2:B3")
    sheet.unmerge_cells(start_row=5, start_column=3, end_row=7, end_column=4)
    # 保存文件
    workbook.save(file_path)




if __name__ == '__main__':
    demo8()





















